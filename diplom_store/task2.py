from openpyxl import load_workbook
import xlwt

list_smeta = ['ИС', 'ССОИ', 'СПДИ', 'СОТН', 'СОПС', 'СКУД', 'ССО', 'СЭОО']

rb = load_workbook('smeta11.xlsx')
# print(rb.get_sheet_names())  # ['02-01-02-искл', '02-01-03-искл', '02-01-04-искл', '02-01-05-искл']
wb = xlwt.Workbook()
for shet_item in rb.get_sheet_names():
    sheet = rb.get_sheet_by_name(shet_item)
    ws = wb.add_sheet(shet_item)

    j = 0
    for i in range(1, sheet.max_row):
        new_row = []
        vid_tovara = sheet.cell(row=i, column=2).value
        # print(vid_tovara)
        if str(vid_tovara).startswith('ТЦ') or str(vid_tovara).startswith('ФССЦ'):
            price = str(sheet.cell(row=i + 2, column=3).value)
            print(price)
            if price.split('/')[0][:4] == 'Цена':
                price = price.split('/')[0][5:]

            if 'Объем' in price or 'Всего по позиции' in price:
                price = str(sheet.cell(row=i, column=10).value)
                # print(price)

            j += 1
            new_row.append(str(sheet.cell(row=i, column=1).value).replace('О', '').rstrip())
            new_row.append(sheet.cell(row=i, column=3).value)
            new_row.append(sheet.cell(row=i, column=6).value)
            new_row.append(sheet.cell(row=i, column=9).value)
            new_row.append(price)
            print(new_row)
            for ind, item in enumerate(new_row):
                ws.write(j - 1, ind, item)
wb.save('new_most.xls')