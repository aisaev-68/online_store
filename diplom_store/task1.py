from openpyxl import load_workbook
import xlwt

list_smeta = ['ИС', 'ССОИ', 'СТН', 'СОТС', 'СКУД', 'ССО', 'СЭ', 'СОО']

rb = load_workbook('lastsmeta.xlsx')
# print(rb.get_sheet_names())  # ['02-01-02-искл', '02-01-03-искл', '02-01-04-искл', '02-01-05-искл']
wb = xlwt.Workbook()
for shet_item in rb.get_sheet_names():
    sheet = rb.get_sheet_by_name(shet_item)
    ws = wb.add_sheet(shet_item)
    # print('99999999', sheet.max_row)

    j = 0
    for i in range(1, sheet.max_row):
        new_row = []
        vid_tovara = sheet.cell(row=i, column=2).value
        if str(vid_tovara).startswith('ТЦ') or str(vid_tovara).startswith('ФССЦ') or str(vid_tovara).startswith(
                'Из состава'):
            price = str(sheet.cell(row=i + 2, column=3).value)
            if price.split('/')[0][:4] == 'Цена':
                price = price.split('/')[0][5:]
            if 'Объем' in price:
                price = str(sheet.cell(row=i + 3, column=3).value)
                price = price.split('/')[0][5:]
            if 'Всего по позиции' in price:
                price = str(sheet.cell(row=i, column=10).value)
            # print(sheet.cell(row=i, column=1).value,
            #       sheet.cell(row=i, column=3).value,
            #       sheet.cell(row=i, column=6).value,
            #       sheet.cell(row=i, column=9).value,
            #       price)
            j += 1
            new_row.append(str(sheet.cell(row=i, column=1).value).replace('О', '').rstrip())
            new_row.append(sheet.cell(row=i, column=3).value)
            new_row.append(sheet.cell(row=i, column=6).value)
            new_row.append(sheet.cell(row=i, column=9).value)
            new_row.append(price)
            for ind, item in enumerate(new_row):
                ws.write(j - 1, ind, item)
wb.save('new03.xls')